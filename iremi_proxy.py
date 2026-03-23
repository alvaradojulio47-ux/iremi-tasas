#!/usr/bin/env python3
"""
IREMI Remesas — Proxy local Binance API
Ejecutar ANTES de abrir el HTML: python iremi_proxy.py
Corre en http://localhost:8765
"""
import hashlib, hmac, time, json, os
import urllib.request, urllib.parse, urllib.error
from http.server import HTTPServer, BaseHTTPRequestHandler

PORT        = 8765
API_KEY     = ''
API_SECRET  = ''
TIME_OFFSET = 0
CREDS_FILE  = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.iremi_creds')

def load_creds():
    """Carga credenciales guardadas localmente"""
    global API_KEY, API_SECRET
    if os.path.exists(CREDS_FILE):
        try:
            with open(CREDS_FILE, 'r') as f:
                lines = f.read().strip().splitlines()
            if len(lines) >= 2:
                API_KEY    = lines[0].strip()
                API_SECRET = lines[1].strip()
                print(f"  ✓ Credenciales cargadas (...{API_KEY[-6:]})")
                return True
        except Exception as e:
            print(f"  ⚠ Error leyendo credenciales: {e}")
    return False

def save_creds():
    """Guarda credenciales en archivo local"""
    try:
        with open(CREDS_FILE, 'w') as f:
            f.write(f"{API_KEY}\n{API_SECRET}\n")
        # Permisos solo para el usuario actual (Unix/Mac)
        try:
            import stat
            os.chmod(CREDS_FILE, stat.S_IRUSR | stat.S_IWUSR)
        except: pass
        print(f"  ✓ Credenciales guardadas en {CREDS_FILE}")
    except Exception as e:
        print(f"  ✗ Error guardando credenciales: {e}")

# ── Sincronización de tiempo ──────────────────────────────
def get_server_time():
    url = 'https://api.binance.com/api/v3/time'
    with urllib.request.urlopen(url, timeout=5) as r:
        return json.loads(r.read())['serverTime']

def sync_time():
    global TIME_OFFSET
    try:
        local_before = int(time.time() * 1000)
        server_time  = get_server_time()
        local_after  = int(time.time() * 1000)
        TIME_OFFSET  = server_time - (local_before + local_after) // 2
        print(f"  ✓ Tiempo sincronizado con Binance (offset: {TIME_OFFSET}ms)")
    except Exception as e:
        print(f"  ⚠ Sync tiempo: {e}")

# ── Firma HMAC-SHA256 ─────────────────────────────────────
def sign(secret, params):
    params['timestamp'] = int(time.time() * 1000) + TIME_OFFSET
    query = urllib.parse.urlencode(params)
    sig   = hmac.new(secret.encode('utf-8'), query.encode('utf-8'), hashlib.sha256).hexdigest()
    return query + '&signature=' + sig

def binance_get(endpoint, params):
    qs  = sign(API_SECRET, params)
    url = f'https://api.binance.com{endpoint}?{qs}'
    req = urllib.request.Request(url, headers={'X-MBX-APIKEY': API_KEY})
    try:
        with urllib.request.urlopen(req, timeout=10) as r:
            return json.loads(r.read().decode())
    except urllib.error.HTTPError as e:
        body = e.read().decode()
        raise Exception(f"Binance {e.code}: {body}")

def p2p_search(fiat, trade_type='SELL', rows=20):
    """
    Busca anuncios P2P en Binance.
    trade_type='SELL' → vendedores de USDT que reciben fiat (para tasas destino)
    trade_type='BUY'  → compradores de USDT que pagan fiat (para tasa CLP)
    Filtra promovidos y retorna precios ordenados.
    """
    url  = 'https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search'
    body = json.dumps({
        "asset":          "USDT",
        "fiat":           fiat,
        "tradeType":      trade_type,   # SELL = ellos venden USDT
        "page":           1,
        "rows":           rows,
        "payTypes":       [],
        "publisherType":  None          # None = incluye todos, luego filtramos promoted
    }).encode()
    req = urllib.request.Request(url, data=body,
        headers={'Content-Type':'application/json','User-Agent':'Mozilla/5.0'})
    with urllib.request.urlopen(req, timeout=10) as r:
        data = json.loads(r.read())
    
    # Filtrar promocionados
    ads = [a for a in data.get('data', [])
           if a.get('adv', {}).get('classify') != 'promoted']
    
    prices = sorted([float(a['adv']['price']) for a in ads], reverse=True)
    return prices   # desc order: [0]=highest, [1]=2nd highest

# ── HTTP Handler ──────────────────────────────────────────
class Handler(BaseHTTPRequestHandler):
    def log_message(self, fmt, *args): pass

    def cors_headers(self, code=200):
        self.send_response(code)
        self.send_header('Access-Control-Allow-Origin',  '*')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.send_header('Access-Control-Allow-Methods', 'GET,POST,OPTIONS')
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.end_headers()

    def send_json(self, data, code=200):
        self.cors_headers(code)
        self.wfile.write(json.dumps(data, ensure_ascii=False).encode('utf-8'))

    def read_body(self):
        n = int(self.headers.get('Content-Length', 0))
        if n:
            try:    return json.loads(self.rfile.read(n))
            except: return {}
        return {}

    def do_OPTIONS(self): self.cors_headers()

    def do_GET(self):
        if self.path == '/ping':
            self.send_json({'ok': True, 'msg': 'Proxy IREMI activo'})
        else:
            self.send_json({'error': 'Not found'}, 404)

    def do_POST(self):
        global API_KEY, API_SECRET
        body = self.read_body()

        # ── /config ───────────────────────────────────────
        if self.path == '/config':
            API_KEY    = body.get('apiKey', '').strip()
            API_SECRET = body.get('secret', '').strip()
            if not API_KEY or not API_SECRET:
                return self.send_json({'error': 'Faltan credenciales'}, 400)
            print(f"  ✓ Credenciales recibidas (key: ...{API_KEY[-6:]})")
            # Guardar si el usuario lo solicita
            if body.get('save', False):
                save_creds()
                print(f"  ✓ Credenciales guardadas permanentemente")
            sync_time()
            self.send_json({'ok': True, 'saved': body.get('save', False)})

        # ── /balance ─────────────────────────────────────
        elif self.path == '/balance':
            if not API_KEY:
                return self.send_json({'error': 'Sin credenciales'}, 401)
            try:
                data  = binance_get('/sapi/v1/capital/config/getall', {})
                usdt  = next((c for c in data if c.get('coin') == 'USDT'), None)
                bal   = [{'asset':'USDT','free': usdt['free'] if usdt else '0'}]
                self.send_json(bal)
                print(f"  ✓ Balance: {usdt['free'] if usdt else '0'} USDT")
            except Exception as e:
                try:
                    data2 = binance_get('/api/v3/account', {})
                    bals  = data2.get('balances', [])
                    usdt2 = next((b for b in bals if b['asset'] == 'USDT'), None)
                    self.send_json([{'asset':'USDT','free': usdt2['free'] if usdt2 else '0'}])
                except Exception as e2:
                    print(f"  ✗ Balance: {e2}")
                    self.send_json({'error': str(e2)}, 500)

        # ── /tasas ────────────────────────────────────────
        # Obtiene TODAS las tasas automáticamente:
        # - P2P Binance: VES, COP, PEN, ARS, MXN, DOP → 2° más alto no promovido
        # - P2P CLP: 2° más barato no promovido (tasa de compra USDT con CLP)
        # - Spot/ER-API: EUR, BOB, USD
        elif self.path == '/tasas':
            tasas  = {}
            detail = {}  # para debug: top 3 precios por par

            # ── P2P pairs SELL (vendedores USDT → reciben fiat) ──
            p2p_fiats = ['VES','COP','PEN','ARS','MXN','DOP']
            for fiat in p2p_fiats:
                try:
                    prices = p2p_search(fiat, 'SELL', rows=20)
                    if len(prices) >= 2:
                        tasas[fiat]  = prices[1]   # 2° más alto no promovido ✓
                        detail[fiat] = prices[:3]
                        print(f"  ✓ USDT/{fiat}: #{1}={prices[0]:,.2f} | #{2}={prices[1]:,.2f} (usado) | #{3}={prices[2] if len(prices)>2 else '—'}")
                    elif len(prices) == 1:
                        tasas[fiat]  = prices[0]
                        detail[fiat] = prices[:3]
                        print(f"  ⚠ USDT/{fiat}: solo 1 precio no promovido → {prices[0]:,.2f}")
                    else:
                        print(f"  ⚠ USDT/{fiat}: sin precios no promovidos")
                except Exception as e:
                    print(f"  ✗ P2P {fiat}: {e}")

            # ── CLP: precio de compra USDT (BUY side, fiat=CLP) ──
            # tradeType=BUY → compradores de USDT que pagan CLP
            # Queremos el 2° más barato (ASC) = mejor precio de compra para IREMI
            try:
                prices_clp = p2p_search('CLP', 'BUY', rows=20)
                prices_clp_asc = sorted(prices_clp)  # ASC: más barato primero
                if len(prices_clp_asc) >= 2:
                    tasas['CLP']   = prices_clp_asc[1]   # 2° más barato
                    detail['CLP']  = prices_clp_asc[:3]
                    print(f"  ✓ USDT/CLP (compra): #{1}=${prices_clp_asc[0]:,.2f} | #{2}=${prices_clp_asc[1]:,.2f} (usado)")
                elif prices_clp_asc:
                    tasas['CLP'] = prices_clp_asc[0]
                    print(f"  ✓ USDT/CLP (compra): ${prices_clp_asc[0]:,.2f} (único disponible)")
            except Exception as e:
                print(f"  ✗ P2P CLP: {e}")

            # ── EUR via Binance bookTicker ─────────────────
            try:
                url_eur = 'https://api.binance.com/api/v3/ticker/bookTicker?symbol=EURUSDT'
                with urllib.request.urlopen(url_eur, timeout=5) as r:
                    d = json.loads(r.read())
                tasas['EUR'] = float(d.get('askPrice', 0))
                print(f"  ✓ EUR (spot): {tasas['EUR']}")
            except Exception as e:
                print(f"  ⚠ EUR spot: {e}")

            # ── BOB, USD via open.er-api fallback ────────
            try:
                er_url = 'https://open.er-api.com/v6/latest/USDT'
                with urllib.request.urlopen(er_url, timeout=6) as r:
                    er = json.loads(r.read())
                er_rates = er.get('rates', {})
                for fiat in ['BOB','USD']:
                    if fiat not in tasas and fiat in er_rates:
                        tasas[fiat] = er_rates[fiat]
                        print(f"  ✓ {fiat} (ER-API): {er_rates[fiat]}")
                # Also fill any P2P gaps
                for fiat in ['COP','PEN','ARS','MXN','DOP','EUR']:
                    if fiat not in tasas and fiat in er_rates:
                        tasas[fiat] = er_rates[fiat]
                        print(f"  ✓ {fiat} (ER-API fallback): {er_rates[fiat]}")
            except Exception as e:
                print(f"  ⚠ ER-API: {e}")

            tasas['USD'] = 1.0  # USD siempre fijo

            print(f"  ✓ Tasas completas: {list(tasas.keys())}")
            self.send_json({'ok': True, 'tasas': tasas, 'detail': detail})

        # ── /p2p ─────────────────────────────────────────
        elif self.path == '/p2p':
            if not API_KEY:
                return self.send_json({'error': 'Sin credenciales'}, 401)
            try:
                params = {
                    'tradeType':      body.get('tradeType', 'SELL'),
                    'startTimestamp': body.get('startTimestamp', int(time.time()*1000)-7*86400*1000),
                    'page':           body.get('page', 1),
                    'rows':           body.get('rows', 100),
                }
                data = binance_get('/sapi/v1/c2c/orderMatch/listUserOrderHistory', params)
                print(f"  ✓ P2P historial {params['tradeType']}: {len(data.get('data',[]))} órdenes")
                self.send_json(data)
            except Exception as e:
                print(f"  ✗ P2P historial: {e}")
                self.send_json({'error': str(e)}, 500)

        # ── /insertar ─────────────────────────────────────
        elif self.path == '/insertar':
            try:
                import os, glob
                from openpyxl import load_workbook
                from datetime import datetime

                mes    = body.get('mes', 'Abril 2026')
                fecha  = body.get('fecha')
                monto  = body.get('monto')
                dc     = body.get('dc')
                dv     = body.get('dv')
                usdt   = body.get('usdt')
                tax    = body.get('tax', -0.06)
                dest   = body.get('dest')
                tiremi = body.get('tiremi')
                coment = body.get('coment')

                script_dir = os.path.dirname(os.path.abspath(__file__))
                excel_path = os.path.join(script_dir, 'IREMI_Sistema_Completo.xlsx')
                if not os.path.exists(excel_path):
                    matches = glob.glob(os.path.join(script_dir,'**','IREMI_Sistema_Completo.xlsx'),recursive=True)
                    if matches: excel_path = matches[0]
                    else: return self.send_json({'error':f'Excel no encontrado en {script_dir}'},404)

                wb = load_workbook(excel_path)
                if mes not in wb.sheetnames:
                    return self.send_json({'error':f"Pestaña '{mes}' no existe"},400)
                ws = wb[mes]

                insert_row = next((r for r in range(2,500) if not ws.cell(r,4).value), None)  # col D=Monto
                if not insert_row:
                    return self.send_json({'error':'Sin fila vacía'},400)

                if fecha:
                    for fmt in ['%Y-%m-%d','%d/%m/%Y','%d-%m-%Y']:
                        try: fecha=datetime.strptime(fecha,fmt); break
                        except: pass

                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

                tipo = 'SELL' if float(usdt or 0) < 0 else 'BUY'
                is_sell = tipo == 'SELL'
                r = insert_row

                # ── Valores manuales (cols A-H, J, L, N) ─────
                ws.cell(r,1).value = r - 1                              # A: N°
                ws.cell(r,2).value = tipo                               # B: Tipo
                ws.cell(r,3).value = fecha                              # C: Fecha
                ws.cell(r,3).number_format = 'DD/MM/YY'
                ws.cell(r,4).value = float(monto) if monto else None    # D: Monto CLP
                ws.cell(r,4).number_format = '$#,##0'
                ws.cell(r,5).value = float(dc) if dc else None          # E: USDT/CLP Compra
                ws.cell(r,5).number_format = '#,##0.00'
                ws.cell(r,6).value = float(dv) if dv else None          # F: USDT/CLP Venta
                ws.cell(r,6).number_format = '#,##0.00'
                ws.cell(r,7).value = float(usdt) if usdt is not None else None  # G: USDT Mov
                ws.cell(r,7).number_format = '#,##0.000'
                ws.cell(r,8).value = -abs(float(tax)) if tax else -0.06 # H: Tax Binance
                ws.cell(r,8).number_format = '#,##0.000'
                if dest:   ws.cell(r,10).value = float(dest)            # J: Tasa Destino
                ws.cell(r,10).number_format = '#,##0.000'
                if tiremi: ws.cell(r,12).value = float(tiremi)          # L: Tasa IREMI
                ws.cell(r,12).number_format = '0.000000'
                if coment: ws.cell(r,14).value = str(coment)            # N: Comentario

                # ── Fórmulas calculadas (cols I, K, M, O, P, Q, R, S) ─
                ws.cell(r,9).value  = f'=I{r-1}+G{r}+H{r}'             # I: Capital USDT
                ws.cell(r,9).number_format = '#,##0.000'
                ws.cell(r,11).value = f'=IF(AND(B{r}="SELL",E{r}>0,J{r}<>""),J{r}/E{r},0)'  # K: Tasa Real
                ws.cell(r,11).number_format = '0.000000'
                ws.cell(r,13).value = f'=IF(AND(B{r}="SELL",L{r}>0),D{r}*L{r},0)'  # M: Valor Destino
                ws.cell(r,13).number_format = '#,##0.00'
                ws.cell(r,15).value = f'=IF(AND(B{r}="SELL",G{r}<>0),Q{r}/ABS(G{r}),0)'  # O: %G USDT
                ws.cell(r,15).number_format = '0.00%'
                ws.cell(r,16).value = f'=IF(AND(B{r}="SELL",D{r}>0),R{r}/D{r},0)'  # P: %G CLP
                ws.cell(r,16).number_format = '0.00%'
                ws.cell(r,17).value = f'=IF(AND(B{r}="SELL",D{r}>0,E{r}>0),(D{r}/E{r})+G{r},0)'  # Q: Gan USDT
                ws.cell(r,17).number_format = '#,##0.000000'
                ws.cell(r,18).value = f'=IF(B{r}="SELL",Q{r}*F{r},0)'  # R: Gan CLP
                ws.cell(r,18).number_format = '$#,##0.00'
                ws.cell(r,19).value = f'=S{r-1}+R{r}'                   # S: Acumulado SELL
                ws.cell(r,19).number_format = '$#,##0.00'

                # ── Formato visual igual al resto de filas ────
                NAVY2='0D1F3C'; DARK='060E1A'; WHITE='FFFFFF'
                SILVER='C8D6E5'; GREEN='4ADE80'; ORANGE='F97316'; GOLD='F5C518'
                bg_color = NAVY2 if (r % 2 == 0) else DARK
                bg = PatternFill('solid', fgColor=bg_color)
                bdr = Border(bottom=Side(border_style='thin', color='1E3256'))
                txt_color = WHITE if is_sell else SILVER

                for col in range(1, 20):
                    c = ws.cell(r, col)
                    c.fill = bg
                    c.alignment = Alignment(horizontal='center', vertical='center')
                    c.border = bdr
                    c.font = Font(name='Arial', size=9, color=txt_color)

                # Colores especiales
                ws.cell(r,2).font  = Font(name='Arial',size=9,color=GREEN if is_sell else ORANGE,bold=True)
                ws.cell(r,18).font = Font(name='Arial',size=9,color=GREEN if is_sell else SILVER,bold=True)
                ws.cell(r,19).font = Font(name='Arial',size=9,color=GOLD,bold=True)

                wb.save(excel_path)
                gan_usdt = (float(monto)/float(dc))+float(usdt) if monto and dc and usdt else 0
                gan_clp  = gan_usdt*float(dv) if dv else 0
                print(f"  ✓ Insertado '{mes}' fila {insert_row} | Tipo={tipo} | Gan: ${gan_clp:,.0f} CLP")
                self.send_json({'ok':True,'fila':insert_row,'mes':mes,
                                'gan_usdt':round(gan_usdt,6),'gan_clp':round(gan_clp,2)})
            except ImportError:
                self.send_json({'error':'openpyxl no instalado. Ejecuta: pip install openpyxl'},500)
            except Exception as e:
                print(f"  ✗ Insertar: {e}")
                self.send_json({'error':str(e)},500)

        # ── /creds_status — verificar si hay credenciales cargadas ──
        elif self.path == '/creds_status':
            has = bool(API_KEY and API_SECRET)
            self.send_json({
                'has_creds':  has,
                'key_suffix': API_KEY[-6:] if has else '',
            })

        # ── /delete_creds — eliminar credenciales guardadas ──
        elif self.path == '/delete_creds':
            try:
                if os.path.exists(CREDS_FILE):
                    os.remove(CREDS_FILE)
                    print("  ✓ Credenciales eliminadas")
                API_KEY = ''; API_SECRET = ''
                self.send_json({'ok': True, 'msg': 'Credenciales eliminadas'})
            except Exception as e:
                self.send_json({'error': str(e)}, 500)

        # ── /borrar — borrar última fila insertada ────────────
        elif self.path == '/borrar':
            try:
                import os, glob
                from openpyxl import load_workbook

                mes   = body.get('mes', 'Abril 2026')
                fila  = body.get('fila')  # fila específica o None = última

                script_dir = os.path.dirname(os.path.abspath(__file__))
                excel_path = os.path.join(script_dir, 'IREMI_Sistema_Completo.xlsx')
                if not os.path.exists(excel_path):
                    matches = glob.glob(os.path.join(script_dir,'**','IREMI_Sistema_Completo.xlsx'),recursive=True)
                    if matches: excel_path = matches[0]
                    else: return self.send_json({'error':'Excel no encontrado'},404)

                wb = load_workbook(excel_path)
                if mes not in wb.sheetnames:
                    return self.send_json({'error':f"Pestaña '{mes}' no existe"},400)
                ws = wb[mes]

                if fila:
                    target_row = int(fila)
                else:
                    # Find last row with data (col D = Monto)
                    target_row = None
                    for r in range(ws.max_row, 1, -1):
                        if ws.cell(r,4).value not in [None,'']:
                            target_row = r
                            break

                if not target_row:
                    return self.send_json({'error':'No hay fila para borrar'},400)

                # Clear the row (preserve formulas in I,K,M,O,P,Q,R,S)
                # Only clear manual input cols: A,B,C,D,E,F,G,H,J,L,N
                for col in [1,2,3,4,5,6,7,8,10,12,14]:
                    ws.cell(target_row,col).value = None

                wb.save(excel_path)
                print(f"  ✓ Borrada fila {target_row} de '{mes}'")
                self.send_json({'ok':True,'fila':target_row,'mes':mes})

            except Exception as e:
                print(f"  ✗ Borrar: {e}")
                self.send_json({'error':str(e)},500)

        else:
            self.send_json({'error':'Ruta no encontrada'},404)

if __name__ == '__main__':
    print("=" * 60)
    print("  IREMI Remesas — Proxy Binance API  v2.0")
    print(f"  Puerto: http://localhost:{PORT}")
    print()
    print("  Rutas disponibles:")
    print("  POST /config    → guardar API Key + Secret")
    print("  POST /tasas     → obtener TODAS las tasas automáticas")
    print("  POST /balance   → saldo USDT en cuenta")
    print("  POST /p2p       → historial órdenes P2P")
    print("  POST /insertar  → insertar operación en Excel")
    print("  GET  /ping      → verificar proxy activo")
    print()
    print("  /tasas obtiene:")
    print("  • VES,COP,PEN,ARS,MXN,DOP → P2P Binance 2° precio no promovido")
    print("  • CLP → P2P Binance 2° precio de compra más barato")
    print("  • EUR → Binance spot bookTicker")
    print("  • BOB,USD → open.er-api.com")
    print()
    # Check if creds file exists
    creds_status = "GUARDADAS" if os.path.exists(CREDS_FILE) else "no guardadas (ingresa en HTML)"
    print(f"  Credenciales: {creds_status}")
    print()
    print("  1. Deja esta ventana abierta")
    print("  2. Abre IREMI_Ingresar_Cambio.html")
    print("  3. Si es primera vez: ingresa API Key, activa 'Guardar'")
    print("  Ctrl+C para detener")
    print("=" * 60)
    # Cargar credenciales guardadas si existen
    if load_creds():
        sync_time()
    else:
        print("  ⚠ Sin credenciales guardadas — ingresa API Key en el HTML")

    try:
        httpd = HTTPServer(('localhost', PORT), Handler)
        httpd.serve_forever()
    except KeyboardInterrupt:
        print("\n✓ Proxy detenido")
    except OSError as e:
        if 'Address already in use' in str(e):
            print(f"\n✗ Puerto {PORT} ya esta en uso")
            print("  Solucion: abre el Administrador de Tareas → busca 'python' → finalizar tarea")
            print("  O reinicia tu computador")
        else:
            print(f"\n✗ Error de red: {e}")
    except Exception as e:
        print(f"\n✗ Error inesperado: {e}")
        import traceback; traceback.print_exc()
    finally:
        # En Windows, evitar que la ventana se cierre instantaneamente
        import sys
        if sys.platform == 'win32':
            print()
            input("Presiona ENTER para cerrar...")
